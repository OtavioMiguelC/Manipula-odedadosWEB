import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
import requests
import unicodedata
from difflib import SequenceMatcher
import os
import json
import io
import zipfile
import math
import re

# =============================================================================
# CONFIGURAÇÕES GERAIS E CONSTANTES
# =============================================================================
st.set_page_config(page_title="Ferramentas Logísticas - Consolida AI", page_icon="📦", layout="wide")

CAMINHO_CACHE_IBGE = 'municipios_ibge_cache.json'
ARQUIVO_MODELO_REGIAO = 'Modelo Região.xlsx'
ARQUIVO_MODELO_ROTA = 'Modelo Rota.xlsx'
ARQUIVO_MODELO_TDE = "Modelo TDE.xlsx"
ARQUIVO_MODELO_CEP = "Modelo CEP.xlsx"

NOME_ABA = 'Base'
COL_CIDADE = 'Destino'
COL_UF = 'UF Destino'
COL_PRAZO = 'Prazo'
COL_IBGE = 'Codigo IBGE'

# =============================================================================
# FUNÇÕES DE APOIO E CACHE
# =============================================================================
def normalizar(texto):
    if pd.isna(texto): return ""
    texto = str(texto).strip().upper()
    try:
        texto_normalizado = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    except:
        return texto 
    texto_normalizado = texto_normalizado.replace("'", " ").replace(".", " ").replace("-", " ")
    while "  " in texto_normalizado:
        texto_normalizado = texto_normalizado.replace("  ", " ")
    return texto_normalizado.strip()

def API_Atualizar_Cache_IBGE():
    try:
        r = requests.get("https://servicodados.ibge.gov.br/api/v1/localidades/municipios", timeout=30, verify=False)
        municipios_api = r.json()
        mapa_final = {}
        for m in municipios_api:
            if not isinstance(m, dict): continue
            nome = normalizar(m.get('nome', ''))
            micro = m.get("microrregiao") or {}
            meso = micro.get("mesorregiao") or {}
            uf_obj = meso.get("UF") or {}
            uf = normalizar(uf_obj.get("sigla", ""))
            if nome and uf: 
                mapa_final[(nome, uf)] = {'nome': nome, 'uf': uf, 'id': m.get('id')}
        lista_dados = list(mapa_final.values())
        with open(CAMINHO_CACHE_IBGE, 'w', encoding='utf-8') as f:
            json.dump(lista_dados, f, ensure_ascii=False, indent=4)
        return True
    except Exception as e:
        st.error(f"Falha ao processar dados IBGE: {e}")
        return False

def gerar_modelo_base_vazio():
    wb = Workbook()
    ws = wb.active
    ws.title = "Base"
    headers = ["Nome da Região", "Destino", "UF Destino", "Prazo", "Codigo IBGE", "DOMINGO", "SEGUNDA", "TERÇA", "QUARTA", "QUINTA", "SEXTA", "SABADO", "FREQUENCIA"]
    ws.append(headers)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def gerar_modelo_cep_vazio():
    if os.path.exists(ARQUIVO_MODELO_CEP):
        with open(ARQUIVO_MODELO_CEP, "rb") as f:
            return f.read()
    wb = Workbook()
    ws = wb.active
    ws.title = "FaixaCEP"
    ws.append([None, None, 'www.lincros.com.br', None, None])
    ws.append([None, None, 'atendimento@lincros.com', None, None])
    ws.append(['Preencha nesta planilha todas as unidades que a transportadora possui.', None, None, None, None])
    ws.append(['ID Localização', 'CEP Inicial', 'CEP Final', 'Nome', 'Ativo'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def limpar_cep(cep_raw):
    if pd.isna(cep_raw):
        return ""
    cep_str = re.sub(r'\D', '', str(cep_raw)).strip()
    if not cep_str:
        return ""
    return cep_str.zfill(8)

def consultar_cep_api(cep_limpo):
    """Consulta cidade e UF por CEP utilizando ViaCEP, AwesomeAPI, ApiCEP e BrasilAPI com fallback inteligente."""
    if not cep_limpo or len(cep_limpo) != 8:
        return None, None, f"CEP Inválido ({cep_limpo})"
    
    if "cache_cep_api" not in st.session_state:
        st.session_state["cache_cep_api"] = {}
    
    cache = st.session_state["cache_cep_api"]
    if cep_limpo in cache and cache[cep_limpo][0] is not None:
        return cache[cep_limpo]

    headers = {'User-Agent': 'Mozilla/5.0'}

    def _tentar_apis(target_cep):
        # 1. ViaCEP
        try:
            url_viacep = f"https://viacep.com.br/ws/{target_cep}/json/"
            resp = requests.get(url_viacep, headers=headers, timeout=3)
            if resp.status_code == 200:
                dados = resp.json()
                if not dados.get("erro"):
                    cidade = dados.get("localidade", "").strip()
                    uf = dados.get("uf", "").strip()
                    if cidade and uf:
                        return cidade, uf, "ViaCEP"
        except Exception:
            pass

        # 2. AwesomeAPI (Excelente para CEPs genéricos de município como 97700000)
        try:
            url_awesome = f"https://cep.awesomeapi.com.br/json/{target_cep}"
            resp = requests.get(url_awesome, headers=headers, timeout=3)
            if resp.status_code == 200:
                dados = resp.json()
                cidade = dados.get("city", "").strip()
                uf = dados.get("state", "").strip()
                if cidade and uf:
                    return cidade, uf, "AwesomeAPI"
        except Exception:
            pass

        # 3. ApiCEP (CDN estático rápido para CEPs municipais)
        try:
            cep_f = f"{target_cep[:5]}-{target_cep[5:]}"
            url_apicep = f"https://cdn.apicep.com/file/apicep/{cep_f}.json"
            resp = requests.get(url_apicep, headers=headers, timeout=3)
            if resp.status_code == 200:
                dados = resp.json()
                if dados.get("ok"):
                    cidade = dados.get("city", "").strip()
                    uf = dados.get("state", "").strip()
                    if cidade and uf:
                        return cidade, uf, "ApiCEP"
        except Exception:
            pass

        # 4. BrasilAPI v1
        try:
            url_b1 = f"https://brasilapi.com.br/api/cep/v1/{target_cep}"
            resp = requests.get(url_b1, headers=headers, timeout=3)
            if resp.status_code == 200:
                dados = resp.json()
                cidade = dados.get("city", "").strip()
                uf = dados.get("state", "").strip()
                if cidade and uf:
                    return cidade, uf, "BrasilAPI"
        except Exception:
            pass

        return None, None, None

    # Tentativa 1: CEP exato informado
    cidade, uf, prov = _tentar_apis(cep_limpo)
    if cidade and uf:
        cache[cep_limpo] = (cidade, uf, prov)
        return cidade, uf, prov

    # Tentativa 2: Se o CEP termina em '000', tentar '001' (para cidades com CEPs por logradouro)
    if cep_limpo.endswith("000"):
        cep_alt = cep_limpo[:5] + "001"
        cidade, uf, prov = _tentar_apis(cep_alt)
        if cidade and uf:
            cache[cep_limpo] = (cidade, uf, prov)
            return cidade, uf, prov

    cache[cep_limpo] = (None, None, "Não Encontrado")
    return None, None, "Não Encontrado"

def processar_modelo_cep(lista_ceps, file_modelo=ARQUIVO_MODELO_CEP, progress_bar=None):
    if os.path.exists(file_modelo):
        wb = load_workbook(file_modelo)
    else:
        wb = Workbook()
        ws_temp = wb.active
        ws_temp.title = "FaixaCEP"
        ws_temp.append([None, None, 'www.lincros.com.br', None, None])
        ws_temp.append([None, None, 'atendimento@lincros.com', None, None])
        ws_temp.append(['Preencha nesta planilha...', None, None, None, None])
        ws_temp.append(['ID Localização', 'CEP Inicial', 'CEP Final', 'Nome', 'Ativo'])

    ws = wb["FaixaCEP"] if "FaixaCEP" in wb.sheetnames else wb.active

    if ws.max_row >= 5:
        ws.delete_rows(5, ws.max_row - 4)

    linha_atual = 5
    resumo_processamento = []
    total = len(lista_ceps)

    for idx, (cep_ini_raw, cep_fim_raw) in enumerate(lista_ceps, start=1):
        if progress_bar:
            progress_bar.progress(idx / total, text=f"Consultando CEP {idx}/{total}...")
        
        cep_ini = limpar_cep(cep_ini_raw)
        cep_fim = limpar_cep(cep_fim_raw) if cep_fim_raw and str(cep_fim_raw).strip() else cep_ini

        if not cep_ini:
            continue

        cidade, uf, status = consultar_cep_api(cep_ini)

        if cidade and uf:
            nome_lincros = f"{cidade} - {uf}"
        else:
            nome_lincros = f"NÃO ENCONTRADO ({cep_ini})"

        ws.cell(row=linha_atual, column=1, value="")
        ws.cell(row=linha_atual, column=2, value=cep_ini)
        ws.cell(row=linha_atual, column=3, value=cep_fim)
        ws.cell(row=linha_atual, column=4, value=nome_lincros)
        ws.cell(row=linha_atual, column=5, value="VERDADEIRO")

        resumo_processamento.append({
            "Linha": idx,
            "CEP Inicial": cep_ini,
            "CEP Final": cep_fim,
            "Cidade": cidade or "-",
            "UF": uf or "-",
            "Nome Lincros": nome_lincros,
            "Status": status
        })

        linha_atual += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, pd.DataFrame(resumo_processamento)

def carregar_lista_cidades_ibge():
    """Lê o cache do IBGE e retorna uma lista formatada para o selectbox."""
    if not os.path.exists(CAMINHO_CACHE_IBGE):
        API_Atualizar_Cache_IBGE()
    try:
        with open(CAMINHO_CACHE_IBGE, 'r', encoding='utf-8') as f:
            dados = json.load(f)
        lista = [f"{m['nome']} - {m['uf']} ({m['id']})" for m in dados]
        return sorted(lista)
    except Exception:
        return []

# =============================================================================
# LÓGICA DE NEGÓCIO E PROCESSAMENTO
# =============================================================================

def processar_ibge(file_or_df):
    if not os.path.exists(CAMINHO_CACHE_IBGE):
        API_Atualizar_Cache_IBGE()
    with open(CAMINHO_CACHE_IBGE, 'r', encoding='utf-8') as f:
        municipios_lista = json.load(f)
    db_ibge_por_uf = {}
    mapa_exato = {} 
    for m in municipios_lista:
        nome_norm = normalizar(m['nome'])
        uf_norm = normalizar(m['uf'])
        id_ibge = m['id']
        mapa_exato[(nome_norm, uf_norm)] = id_ibge
        if uf_norm not in db_ibge_por_uf: db_ibge_por_uf[uf_norm] = []
        db_ibge_por_uf[uf_norm].append({'nome_norm': nome_norm, 'id': id_ibge, 'nome_real': m['nome']})
    
    if isinstance(file_or_df, pd.DataFrame):
        df = file_or_df.copy()
        is_df = True
    else:
        df = pd.read_excel(file_or_df, sheet_name=NOME_ABA)
        file_or_df.seek(0)
        is_df = False

    if COL_IBGE not in df.columns: df[COL_IBGE] = ""
    count_exato, count_aprox = 0, 0
    nao_encontrados = []
    
    ibge_novos = []
    for index, row in df.iterrows():
        cidade_excel_raw = str(row[COL_CIDADE])
        uf_excel_raw = str(row[COL_UF])
        cidade_norm = normalizar(cidade_excel_raw)
        uf_norm = normalizar(uf_excel_raw)
        ibge_encontrado = mapa_exato.get((cidade_norm, uf_norm))
        if ibge_encontrado:
            count_exato += 1
        else:
            if uf_norm in db_ibge_por_uf:
                candidatos = db_ibge_por_uf[uf_norm]
                melhor_ratio = 0.0
                melhor_candidato = None
                for item in candidatos:
                    ratio = SequenceMatcher(None, cidade_norm, item['nome_norm']).ratio()
                    if ratio > melhor_ratio:
                        melhor_ratio = ratio
                        melhor_candidato = item
                if melhor_ratio >= 0.80 and melhor_candidato:
                    ibge_encontrado = melhor_candidato['id']
                    count_aprox += 1
        if ibge_encontrado:
            ibge_novos.append(ibge_encontrado)
        else:
            ibge_novos.append(row.get(COL_IBGE, ""))
            nao_encontrados.append(f"{cidade_excel_raw} - {uf_excel_raw}")

    df[COL_IBGE] = ibge_novos

    if is_df:
        return df, count_exato, count_aprox, nao_encontrados

    wb = load_workbook(file_or_df)
    ws = wb[NOME_ABA]
    col_ibge_num = df.columns.get_loc(COL_IBGE) + 1 
    for index, val in enumerate(ibge_novos):
        if val: ws.cell(row=index + 2, column=col_ibge_num).value = val

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, count_exato, count_aprox, nao_encontrados

def processar_prazos(file_destino, file_base):
    df_base = pd.read_excel(file_base, sheet_name="Base", dtype=str).fillna("")
    df_base.columns = [str(col).strip().upper() for col in df_base.columns]
    nome_coluna_texto_freq = None
    for col in df_base.columns:
        amostra = df_base[col].head(20).astype(str).tolist()
        for val in amostra:
            if "....." in val or "STQQS" in val:
                nome_coluna_texto_freq = col; break
        if nome_coluna_texto_freq: break
    mapa_colunas_base = {}
    padroes_busca = {'Seg': 'SEG', 'Ter': 'TER', 'Qua': 'QUA', 'Qui': 'QUI', 'Sex': 'SEX', 'Sáb': 'SAB', 'Dom': 'DOM'}
    for dia_destino, texto_busca in padroes_busca.items():
        for col_real in df_base.columns:
            if texto_busca in col_real:
                mapa_colunas_base[dia_destino] = col_real; break
    col_ibge_nome = next((c for c in df_base.columns if 'IBGE' in c), 'CODIGO IBGE')
    df_base = df_base.dropna(subset=[col_ibge_nome])
    df_base['IBGE_LIMPO'] = df_base[col_ibge_nome].apply(lambda x: str(x).split('.')[0].strip())
    df_base = df_base.drop_duplicates(subset=['IBGE_LIMPO'], keep='first')
    dicionario_base = df_base.set_index('IBGE_LIMPO').to_dict('index')
    wb = load_workbook(file_destino)
    nome_aba_destino = "Prazo (localizações)"
    sheet = wb[nome_aba_destino] if nome_aba_destino in wb.sheetnames else wb.active
    header_row = 4
    col_map_dest = {}
    if sheet.max_row >= header_row:
        for cell in sheet[header_row]:
            if cell.value: col_map_dest[str(cell.value).strip()] = cell.column
    else: raise Exception(f"A planilha de destino não tem dados na linha {header_row}.")
    idx_ibge = col_map_dest.get('Código IBGE da Cidade')
    if not idx_ibge:
        for k, v in col_map_dest.items():
            if "IBGE" in k.upper() and "CIDADE" in k.upper():
                idx_ibge = v; break
    if not idx_ibge: raise Exception("Coluna de IBGE não encontrada no destino.")
    cidades_atualizadas = 0
    for row_index in range(header_row + 1, sheet.max_row + 1):
        cell_ibge = sheet.cell(row=row_index, column=idx_ibge)
        if not cell_ibge.value: continue
        ibge_chave = str(cell_ibge.value).split('.')[0].strip()
        if ibge_chave in dicionario_base:
            dados_linha = dicionario_base[ibge_chave]
            if 'Prazo' in col_map_dest and 'PRAZO' in dados_linha:
                sheet.cell(row=row_index, column=col_map_dest['Prazo']).value = dados_linha['PRAZO']
            eh_caso_pontinhos = False
            if nome_coluna_texto_freq:
                texto_freq = str(dados_linha.get(nome_coluna_texto_freq, "")).strip()
                if "......" in texto_freq or (len(texto_freq) > 3 and set(texto_freq) == {'.'}): eh_caso_pontinhos = True
            if eh_caso_pontinhos:
                for dia in ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb']:
                    if dia in col_map_dest: sheet.cell(row=row_index, column=col_map_dest[dia]).value = 'VERDADEIRO'
                if 'Dom' in col_map_dest: sheet.cell(row=row_index, column=col_map_dest['Dom']).value = 'FALSO'
            else:
                for dia_curto, nome_coluna_base in mapa_colunas_base.items():
                    if dia_curto in col_map_dest:
                        valor_bruto = str(dados_linha.get(nome_coluna_base, "")).upper().strip()
                        eh_verdadeiro = valor_bruto in ['VERDADEIRO', 'TRUE', 'S', 'SIM', '1', 'X']
                        sheet.cell(row=row_index, column=col_map_dest[dia_curto]).value = 'VERDADEIRO' if eh_verdadeiro else 'FALSO'
            cidades_atualizadas += 1
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, cidades_atualizadas

def processar_regiao(cnpj, file_or_df, file_modelo):
    if isinstance(file_or_df, pd.DataFrame):
        df_prazos = file_or_df.copy()
    else:
        df_prazos = pd.read_excel(file_or_df, sheet_name='Base')
        
    df_prazos['Nome da Região'] = df_prazos['Nome da Região'].astype(str).str.strip()
    df_prazos = df_prazos[df_prazos['Nome da Região'].notna() & (df_prazos['Nome da Região'].str.upper() != 'NAN') & (df_prazos['Nome da Região'] != '')]
    
    df_prazos['NomeRegiao'] = df_prazos['Nome da Região'].str.upper()
    wb_modelo = load_workbook(file_modelo)
    ws_regioes = wb_modelo['regioes']; ws_localizacoes = wb_modelo['localizacoes_atendidas']
    for ws in [ws_regioes, ws_localizacoes]:
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row: cell.value = None
    for i, nome_regiao in enumerate(df_prazos['NomeRegiao'].unique(), start=5):
        ws_regioes[f'B{i}'] = cnpj; ws_regioes[f'C{i}'] = nome_regiao; ws_regioes[f'D{i}'] = "VERDADEIRO"
    for i, row in enumerate(df_prazos.iterrows(), start=5):
        ws_localizacoes[f'B{i}'] = row[1]['NomeRegiao']; ws_localizacoes[f'E{i}'] = row[1]['Codigo IBGE']
    output = io.BytesIO()
    wb_modelo.save(output)
    output.seek(0)
    return output

def processar_rotas(escolha_rota, cnpj_transportadora, nome_transportadora, desc_adicional, tipo_origem, valor_origem, file_modelo_regioes, file_template_rota):
    wb_modelo_regioes = load_workbook(file_modelo_regioes)
    ws_regioes = wb_modelo_regioes['regioes']
    
    regioes_encontradas = [str(ws_regioes.cell(row=i, column=3).value) for i in range(5, ws_regioes.max_row + 1) if ws_regioes.cell(row=i, column=3).value]
    if not regioes_encontradas: 
        raise Exception("Nenhuma região encontrada no modelo de regiões.")
        
    wb_rotas = load_workbook(file_template_rota)
    ws_rotas = wb_rotas["Rotas"] if "Rotas" in wb_rotas.sheetnames else wb_rotas.active
    
    next_row = 6
    while ws_rotas.cell(row=next_row, column=1).value is not None: 
        next_row += 1
        
    for regiao_destino in regioes_encontradas:
        ws_rotas.cell(row=next_row, column=1).value = f"{cnpj_transportadora} - {nome_transportadora}"
        
        desc = f"{desc_adicional} x {regiao_destino}" if desc_adicional else regiao_destino
        ws_rotas.cell(row=next_row, column=2).value = desc
        
        if tipo_origem == "Cidade (IBGE)": 
            ws_rotas.cell(row=next_row, column=3).value = valor_origem
        else: 
            ws_rotas.cell(row=next_row, column=5).value = valor_origem
            
        ws_rotas.cell(row=next_row, column=8).value = regiao_destino
        ws_rotas.cell(row=next_row, column=10).value = "VERDADEIRO"
        ws_rotas.cell(row=next_row, column=11).value = "VERDADEIRO"
        next_row += 1
        
    output = io.BytesIO()
    wb_rotas.save(output)
    output.seek(0)
    return output

def converter_freq(file):
    wb = load_workbook(file); ws = wb[NOME_ABA]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=12):
        for cell in row:
            valor = str(cell.value).strip().upper() if cell.value is not None else ""
            if valor == "S": cell.value = "VERDADEIRO"
            elif valor == "N": cell.value = "FALSO"
    output = io.BytesIO(); wb.save(output); output.seek(0); return output

def converter_freq_txt(file):
    wb = load_workbook(file); ws = wb.active
    colunas_destino = [7, 8, 9, 10, 11, 12]; letras_referencia = ['S', 'T', 'Q', 'Q', 'S', 'S']
    for row in ws.iter_rows(min_row=2):
        texto_raw = str(row[12].value or "").strip().upper()
        for i, col_idx in enumerate(colunas_destino):
            if i < len(texto_raw) and texto_raw[i] == letras_referencia[i]: row[col_idx - 1].value = True
            else: row[col_idx - 1].value = False
        row[5].value = False
    output = io.BytesIO(); wb.save(output); output.seek(0); return output

def gerar_restricoes_zip(texto_input, template_bytes, limite_linhas, categoria, tipo_f_j, usar_valor):
    linhas = [l for l in texto_input.split('\n') if l.strip()]
    grupos_por_valor = {}

    for l in linhas:
        partes = l.replace('\t', ' ').strip().split()
        if len(partes) < 2: continue 
        
        cnpj_raw = partes[0].strip().replace('.', '').replace('/', '').replace('-', '')
        possivel_valor = partes[-1].replace("R$", "").replace(".", "").replace(",", ".")
        valor_final = "0"
        razao_partes = partes[1:]

        try:
            float(possivel_valor)
            valor_final = possivel_valor
            razao_partes = partes[1:-1]
        except:
            valor_final = "0" 

        razao = " ".join(razao_partes).upper()
        chave_grupo = valor_final if usar_valor else "GERAL"

        if chave_grupo not in grupos_por_valor:
            grupos_por_valor[chave_grupo] = []
        grupos_por_valor[chave_grupo].append((cnpj_raw, razao, valor_final))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for valor_chave, dados_lista in grupos_por_valor.items():
            total_registros = len(dados_lista)
            num_arquivos = math.ceil(total_registros / limite_linhas)
            
            for i in range(num_arquivos):
                inicio = i * limite_linhas
                fim = min((i + 1) * limite_linhas, total_registros)
                lote = dados_lista[inicio:fim]
                
                template_bytes.seek(0)
                wb = load_workbook(template_bytes)
                ws = wb["Pessoa"] if "Pessoa" in wb.sheetnames else wb.active

                row_idx = 5
                for cnpj, razao, v_lin in lote:
                    ws.cell(row=row_idx, column=1).value = str(cnpj)
                    ws.cell(row=row_idx, column=2).value = str(tipo_f_j) 
                    ws.cell(row=row_idx, column=4).value = str(razao)
                    row_idx += 1

                prefixo = f"{categoria} " if categoria != "Outros" else ""
                faixa_linhas = f"{inicio} a {fim} Linhas"
                sufixo_valor = f" R${valor_chave}" if (usar_valor and valor_chave != "0") else ""
                
                nome_arquivo = f"{prefixo}{faixa_linhas}{sufixo_valor}.xlsx".strip()
                
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                zip_file.writestr(nome_arquivo, excel_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer



# =============================================================================
# INTERFACE DO STREAMLIT
# =============================================================================

st.title("📦 Ferramentas Logísticas")

with st.sidebar:
    st.header("⚙️ Configurações Padrão")
    st.download_button(label="📥 Baixar Modelo Base (Vazio)", data=gerar_modelo_base_vazio(), file_name="Base_de_Origem_Template.xlsx", use_container_width=True)
    st.download_button(label="📥 Baixar Modelo CEP (Vazio)", data=gerar_modelo_cep_vazio(), file_name="Modelo CEP.xlsx", use_container_width=True)
    st.divider()
    cnpj_global = st.text_input("CNPJ Transportadora Padrão", value="Preencher aqui")
    nome_global = st.text_input("Nome Transportadora Padrão", value="Preencher aqui")
    
    st.divider()
    if st.button("Atualizar Cache IBGE", use_container_width=True):
        API_Atualizar_Cache_IBGE()
        st.success("Cache IBGE atualizado com sucesso!")

# Abas do aplicativo
tab_cep, tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📮 Cadastro CEP",
    "🌍 Preencher IBGE", 
    "⏱️ Prazos/Freq", 
    "🗺️ Criar Região", 
    "📍 Gerar Rotas", 
    "🔄 Conv. S/N", 
    "📅 Conv. STQQS", 
    "👥 Restrições Por Pessoas"
])

# --- ABA CEPS: CADASTRO CEP LINCROS ---
with tab_cep:
    st.markdown("### 📮 Cadastro CEP (Lincros)")
    st.info("Cole os CEPs iniciais e finais nas colunas abaixo (aceita com ou sem hífen `-`). O sistema consulta automaticamente a cidade/UF via API dos Correios (ViaCEP/BrasilAPI), insere no formato `Cidade - UF` na coluna Nome, mantendo o ID de Localização vazio e o status Ativo como VERDADEIRO.")

    col_input1, col_input2 = st.columns(2)
    txt_cep_ini = col_input1.text_area(
        "CEP Inicial (um por linha)",
        height=250,
        placeholder="Ex:\n80000-000\n97700-000\n01001-000",
        key="txt_cep_ini"
    )
    txt_cep_fim = col_input2.text_area(
        "CEP Final (um por linha - opcional)",
        height=250,
        placeholder="Ex:\n80000-000\n97700-000\n01001-000\n(Se deixar em branco, usará o CEP Inicial)",
        key="txt_cep_fim"
    )

    st.markdown("#### 📁 Ou faça upload de uma planilha contendo CEPs (opcional)")
    st.caption("💡 **Como funciona o upload:** Aceita qualquer planilha Excel (`.xlsx` / `.xls`) ou `.csv`. O sistema detecta automaticamente colunas com nomes como `CEP`, `CEP Inicial` ou `CEP Final`. Caso a planilha não possua cabeçalho, ele utilizará a 1ª coluna como CEP Inicial e a 2ª como CEP Final.")
    file_cep_up = st.file_uploader("Upload de arquivo Excel/CSV", type=["xlsx", "xls", "csv"], key="file_cep_up")

    col_btn_cep1, col_btn_cep2 = st.columns([2, 1])
    if col_btn_cep1.button("🚀 PROCESSAR CEPS & CONSULTAR CORREIOS", use_container_width=True):
        lista_pares = []

        # 1. Obter dados das text areas
        linhas_ini = [l.strip() for l in txt_cep_ini.split('\n') if l.strip()]
        linhas_fim = [l.strip() for l in txt_cep_fim.split('\n') if l.strip()]

        for i, c_ini in enumerate(linhas_ini):
            c_fim = linhas_fim[i] if i < len(linhas_fim) else c_ini
            lista_pares.append((c_ini, c_fim))

        # 2. Obter dados do upload se enviado
        if file_cep_up is not None:
            try:
                ext = file_cep_up.name.split('.')[-1].lower()
                if ext in ['xlsx', 'xls']:
                    df_up = pd.read_excel(file_cep_up)
                else:
                    df_up = pd.read_csv(file_cep_up)

                cols = [str(c).upper() for c in df_up.columns]
                col_ini_idx = next((i for i, c in enumerate(cols) if 'INI' in c or 'CEP' in c), 0)
                col_fim_idx = next((i for i, c in enumerate(cols) if 'FIM' in c or 'FINAL' in c), None)

                for _, row in df_up.iterrows():
                    val_ini = str(row.iloc[col_ini_idx]) if len(row) > col_ini_idx else ""
                    val_fim = str(row.iloc[col_fim_idx]) if col_fim_idx is not None and len(row) > col_fim_idx else val_ini
                    if val_ini and val_ini.lower() != "nan":
                        lista_pares.append((val_ini, val_fim))
            except Exception as e_up:
                st.error(f"Erro ao ler arquivo de CEPs: {e_up}")

        if not lista_pares:
            st.warning("⚠️ Cole os CEPs nas caixas de texto acima ou envie um arquivo para processar.")
        else:
            p_bar = st.progress(0, text="Iniciando consulta de CEPs...")
            try:
                out_cep_bytes, df_resumo = processar_modelo_cep(lista_pares, file_modelo=ARQUIVO_MODELO_CEP, progress_bar=p_bar)
                st.session_state['out_modelo_cep'] = out_cep_bytes
                st.session_state['df_resumo_cep'] = df_resumo
                p_bar.progress(1.0, text="Concluído!")

                sucessos = len(df_resumo[df_resumo['Status'] != 'Não Encontrado'])
                st.success(f"✅ Processamento concluído! {sucessos}/{len(df_resumo)} CEPs localizados com sucesso.")
            except Exception as e_proc:
                st.error(f"Erro durante o processamento: {e_proc}")

    if col_btn_cep2.button("📥 Baixar Modelo CEP (Vazio)", use_container_width=True):
        st.download_button(
            label="📥 Confirmar Download Modelo Vazio",
            data=gerar_modelo_cep_vazio(),
            file_name="Modelo CEP.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    if 'df_resumo_cep' in st.session_state and not st.session_state['df_resumo_cep'].empty:
        st.markdown("#### 📋 Pré-visualização dos Dados")
        st.dataframe(st.session_state['df_resumo_cep'], use_container_width=True)

    if 'out_modelo_cep' in st.session_state:
        st.download_button(
            label="📥 BAIXAR PLANILHA CADASTRO CEP PREENCHIDA (.XLSX)",
            data=st.session_state['out_modelo_cep'],
            file_name="Cadastro CEP Preenchido.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )



# --- ABA 1: IBGE ---
with tab1:
    st.markdown("### Preencher Códigos IBGE")
    file_ibge = st.file_uploader("Planilha de Base", type=["xlsx"], key="ibge_file")
    if file_ibge and st.button("Processar IBGE"):
        out_bytes, exatos, aprox, nao_enc = processar_ibge(file_ibge)
        st.session_state['out_ibge'] = out_bytes
        st.success(f"✅ Sucesso! Exatos: {exatos} | IA: {aprox}")
    if 'out_ibge' in st.session_state:
        st.download_button("📥 Baixar Arquivo IBGE", data=st.session_state['out_ibge'], file_name="Base_IBGE_Preenchida.xlsx")

# --- ABA 2: Prazos e Frequência ---
with tab2:
    st.markdown("### Preencher Prazos e Frequência")
    c1, c2 = st.columns(2); f_dest = c1.file_uploader("Planilha DESTINO", type=["xlsx"]); f_base = c2.file_uploader("BASE", type=["xlsx"])
    if f_dest and f_base and st.button("Processar Prazos"):
        out, at = processar_prazos(f_dest, f_base)
        st.session_state['out_prazos'] = out; st.success(f"✅ {at} cidades atualizadas!")
    if 'out_prazos' in st.session_state:
        st.download_button("📥 Baixar Destino", data=st.session_state['out_prazos'], file_name="Destino_Prazos.xlsx")

# --- ABA 3: Criar Regiões ---
with tab3:
    st.markdown("### Criar Regiões")
    f_reg = st.file_uploader("Base de Prazos", type=["xlsx"], key="reg_up")
    if f_reg and st.button("Criar Regiões"):
        if not cnpj_global: st.warning("Preencha o CNPJ lateral")
        else:
            out = processar_regiao(cnpj_global, f_reg, ARQUIVO_MODELO_REGIAO)
            st.session_state['out_regiao'] = out
            st.success("✅ Regiões criadas!")
    if 'out_regiao' in st.session_state:
        st.download_button("📥 Baixar Regiões", data=st.session_state['out_regiao'], file_name=f"Regioes_{nome_global}.xlsx")

# --- ABA 4: Gerar Rotas ---
with tab4:
    st.markdown("### Gerar Rotas")
    
    file_modelo_regioes = st.file_uploader("1. Modelo de Região (Já preenchido no passo anterior)", type=["xlsx"], key="rota_mod_reg")
    
    st.divider()
    
    col1, col2 = st.columns(2)
    tipo_rota = col1.selectbox("Dados da Rota", ["1: ROTA - PRAZO", "2: (TRANS) (ORIGEM)"])
    cnpj_rota = col2.text_input("CNPJ Transportadora (se difere do padrão)", value=cnpj_global)
    nome_transp_rota = col1.text_input("Nome Transportadora", value=nome_global) 
    
    desc_rota = col2.text_input("Desc. Adicional (Opcional)")
    
    st.markdown("#### Dados de Origem")
    col3, col4 = st.columns(2)
    tipo_origem = col3.radio("Definir origem por:", ["Cidade (IBGE)", "Região"])
    
    if tipo_origem == "Cidade (IBGE)":
        lista_cidades = carregar_lista_cidades_ibge()
        cidade_selecionada = col4.selectbox("Selecione ou digite a Cidade de Origem", options=[""] + lista_cidades)
        
        if cidade_selecionada:
            valor_origem = cidade_selecionada.split("(")[-1].replace(")", "").strip()
        else:
            valor_origem = ""
    else:
        valor_origem = col4.text_input("Nome da Região de Origem")
    
    if file_modelo_regioes and st.button("Gerar Rotas"):
        if not cnpj_rota or not valor_origem:
            st.warning("⚠️ CNPJ e Origem são obrigatórios.")
        elif not os.path.exists(ARQUIVO_MODELO_ROTA):
            st.error(f"⚠️ O arquivo original '{ARQUIVO_MODELO_ROTA}' não foi encontrado na pasta do projeto!")
        else:
            with st.spinner(f"Gerando rotas baseadas no banco ({ARQUIVO_MODELO_ROTA})..."):
                try:
                    out_bytes = processar_rotas(
                        tipo_rota.split(":")[0], 
                        cnpj_rota, 
                        nome_transp_rota.upper(), 
                        desc_rota.upper(), 
                        tipo_origem, 
                        valor_origem, 
                        file_modelo_regioes, 
                        ARQUIVO_MODELO_ROTA
                    )
                    st.session_state['out_rotas'] = out_bytes
                    
                    nome_sugerido_rota = f"Rotas_{nome_transp_rota.strip()}.xlsx" if nome_transp_rota.strip() else "Rotas_Preenchidas.xlsx"
                    st.session_state['nome_arq_rota'] = nome_sugerido_rota
                    
                    st.success("✅ Rotas estruturadas com sucesso dentro do modelo original!")
                except Exception as e:
                    st.error(f"Erro: {e}")
                    
    if 'out_rotas' in st.session_state:
        st.download_button(
            label="📥 Baixar Planilha de Rotas Preenchida", 
            data=st.session_state['out_rotas'], 
            file_name=st.session_state.get('nome_arq_rota', 'Rotas_Preenchidas.xlsx'), 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# --- ABA 5: Converter S/N ---
with tab5:
    f_sn = st.file_uploader("Planilha S/N", type=["xlsx"])
    if f_sn and st.button("Converter S/N"):
        st.session_state['out_sn'] = converter_freq(f_sn); st.success("Convertido!")
    if 'out_sn' in st.session_state:
        st.download_button("📥 Baixar S/N", data=st.session_state['out_sn'], file_name="S_N_Convertido.xlsx")

# --- ABA 6: Converter STQQS ---
with tab6:
    f_st = st.file_uploader("Planilha STQQS", type=["xlsx"])
    if f_st and st.button("Converter STQQS"):
        st.session_state['out_stqqs'] = converter_freq_txt(f_st); st.success("Convertido!")
    if 'out_stqqs' in st.session_state:
        st.download_button("📥 Baixar STQQS", data=st.session_state['out_stqqs'], file_name="STQQS_Convertido.xlsx")

# --- ABA 7: RESTRIÇÕES POR PESSOAS ---
with tab7:
    st.markdown("### 👥 Restrições Por Pessoas")
    st.info("Gera arquivos de cadastro de pessoas separando por valor e limite de linhas.")

    col_btn1, col_btn2, col_btn3 = st.columns(3)
    categoria_fleg = col_btn1.radio("Selecione a Categoria:", ["TDE", "TAE", "Outros"], horizontal=True)
    tipo_pessoa_fleg = col_btn2.radio("Tipo de Pessoa:", ["J", "F"], horizontal=True)
    usar_valor_fleg = col_btn3.checkbox("Filtrar/Incluir Valor no nome?", value=True)

    limite_linhas_rest = st.number_input("Linhas por arquivo:", min_value=1, value=500, step=100)
    
    st.markdown("**📋 Cole abaixo os dados (Formato: CNPJ | Razão Social | Valor Opcional):**")
    texto_rest = st.text_area("Dados:", height=250, placeholder="Ex: 12345678000100 EMPRESA LTDA 250,00", key="txt_rest")

    if st.button("🚀 PROCESSAR RESTRIÇÕES POR PESSOAS", use_container_width=True):
        if not os.path.exists(ARQUIVO_MODELO_TDE):
            st.error(f"Arquivo '{ARQUIVO_MODELO_TDE}' não encontrado na pasta do projeto!")
        elif not texto_rest.strip():
            st.warning("Cole os dados para processar.")
        else:
            with st.spinner("Estruturando dados e gerando arquivos..."):
                try:
                    with open(ARQUIVO_MODELO_TDE, "rb") as f:
                        template_bytes = io.BytesIO(f.read())
                    
                    zip_out = gerar_restricoes_zip(
                        texto_rest, 
                        template_bytes, 
                        int(limite_linhas_rest), 
                        categoria_fleg, 
                        tipo_pessoa_fleg, 
                        usar_valor_fleg
                    )
                    st.session_state['zip_rest'] = zip_out
                    st.success("✅ Arquivos gerados e compactados com sucesso!")
                except Exception as e:
                    st.error(f"Erro: {e}")

    if 'zip_rest' in st.session_state:
        st.download_button(
            label="📥 BAIXAR ARQUIVOS GERADOS (ZIP)",
            data=st.session_state['zip_rest'],
            file_name=f"Restricoes_{categoria_fleg}.zip",
            mime="application/zip",
            use_container_width=True
        )
